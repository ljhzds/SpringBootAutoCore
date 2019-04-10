# coding: utf-8
from collections import namedtuple
from openpyxl import load_workbook

from .utils import columnname_to_valname

Table = namedtuple('Table', ['name', 'cn_name', 'dao_namespace', 'model_class'])
Column = namedtuple('Column',
                    ['index', 'name', 'cn_name', 'data_type', 'length', 'is_key', 'is_allow_null', 'description',
                     'java_data_type', 'java_name'])


def read_app_config(desc_excel_file_path):
    config_dict = {}
    try:
        wb = load_workbook(filename=desc_excel_file_path, data_only=True)
        sheet = wb.worksheets[0]
        rows = list(sheet.rows)

        for row in rows:
            k = row[1].value
            v = row[2].value
            if k and v:
                config_dict.update({k: v})
    except Exception as e:
        print(e)
        pass
    return config_dict


def get_tables(desc_excel_file_path):
    tables = []

    wb = load_workbook(filename=desc_excel_file_path, data_only=True)

    for sheet in wb.worksheets[1:]:
        try:
            rows = list(sheet.rows)
            table_info_row = rows[0]
            table = Table(table_info_row[1].value, table_info_row[3].value, table_info_row[5].value,
                          table_info_row[7].value)
            column_list = []
            for row in rows[2:]:
                data_type = row[3].value.upper()
                if data_type == 'VARCHAR':
                    java_data_type = 'String'
                elif data_type == 'BIGINT':
                    java_data_type = 'long'
                elif data_type == 'INT':
                    java_data_type = 'int'
                elif data_type == 'FLOAT':
                    java_data_type = 'float'
                elif data_type == 'DOUBLE':
                    java_data_type = 'Double'
                elif data_type == 'DECIMAL' or data_type == 'NUMERIC':
                    java_data_type = 'BigDecimal'
                else:
                    java_data_type = data_type.lower()
                java_name = columnname_to_valname(row[1].value)
                column = Column(*[c.value for c in row], java_data_type, java_name)
                column_list.append(column)
            tables.append((table, column_list))
        except Exception as e:
            print(e)
            continue

    return tables


if __name__ == '__main__':
    tables = get_tables('C:\\Users\\13676\\PycharmProjects\\pyDailyScripts\\spring_helper\\templates\\tables.xlsx')
    table, column_list = tables[0]
    print(table.cn_name, len(column_list))
